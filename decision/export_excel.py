"""
將策略分析結果匯出為 Excel（含背景前提、決策結論、傳統 vs ILP 比較）。
"""
from __future__ import annotations

from datetime import datetime
from pathlib import Path
from typing import Optional

import pandas as pd

from config_columns import COST_BY_PROFILE, COST_TIERS, DEFAULT_PERIODS, N_SYNTHETIC_SONGS
from paths import OUTPUT_DIR, STRATEGY_EXCEL, SYNTHETIC_CSV


def _background_rows(
    csv_path: Path,
    profile: dict,
    cfg: dict,
    cv_note: str = '',
) -> list[tuple[str, str]]:
    return [
        ('專案名稱', 'Hit Song Intelligence：占卜出下一首爆款神曲'),
        ('報告類型', '20 首模擬新歌 · 三檔期 · 多限制式背包決策'),
        ('產出時間', datetime.now().strftime('%Y-%m-%d %H:%M')),
        ('', ''),
        ('一、研究背景', ''),
        (
            '產業情境',
            '傳統 A&R 各檔期獨立依 ŷ 貪婪選曲，常忽略跨檔預算、發行量上下限與檔期平均聲學門檻，'
            '導致組合不可行或總熱度次優。',
        ),
        (
            '資料來源',
            f'MSD 預處理樣本：{csv_path.name}（{profile.get("n_rows", "")} 列）。',
        ),
        (
            '決策設計',
            f'1 首原型 → {cfg.get("n_songs", N_SYNTHETIC_SONGS)} 首假新歌；'
            f'ILP 納入 PDF §4.2 式 (1)–(7)：預算、檔期 L/U、聲學平均、每曲僅排一檔。',
        ),
        ('', ''),
        ('二、方法', ''),
        ('目標變數', f'{profile.get("target", "song_hotttnesss")}（0–1）'),
        ('預測模型', 'HistGradientBoostingRegressor' + (f' {cv_note}' if cv_note else '')),
        (
            '成本設定',
            '每首依情境 profile 自訂成本（14–60），見 config_columns.COST_BY_PROFILE；'
            f'FLOP/HOP/TOP 分級門檻 ŷ≥{COST_TIERS[1]["y_min"]}/{COST_TIERS[0]["y_min"]} 僅供傳統對照。',
        ),
        (
            '檔期',
            '、'.join(cfg.get('periods', DEFAULT_PERIODS)),
        ),
        ('', ''),
        ('三、聲明', ''),
        ('預測限制', 'head200 小樣本；ŷ 為決策輔助指標，非播放量保證。'),
    ]


def _decision_rows(
    chosen: list[str],
    kparams: dict,
    summary_row: dict,
    ranked: pd.DataFrame,
    multperiod_df: Optional[pd.DataFrame],
    comparison: Optional[pd.DataFrame],
    feas_note: Optional[str],
) -> list[tuple[str, str]]:
    lines = [
        ('決策標題', '三檔期發行組合（ILP 最優解）'),
        ('求解狀態', str(summary_row.get('knapsack_status', ''))),
        ('總預算（三期加總）', str(summary_row.get('total_budget', ''))),
        ('ILP Σŷ / 成本', f'{summary_row.get("total_y_hat_ilp", 0):.4f} / {summary_row.get("total_cost_ilp", 0):.0f}'),
        ('傳統 Σŷ / 成本', f'{summary_row.get("total_y_hat_traditional", 0):.4f} / {summary_row.get("total_cost_traditional", 0):.0f}'),
        ('Δŷ（ILP − 傳統）', str(summary_row.get('delta_yhat_ilp_vs_traditional', ''))),
        ('', ''),
    ]
    if feas_note:
        lines.append(('備註', feas_note))
    lines.append(('選中歌曲', '、'.join(chosen)))
    if multperiod_df is not None and len(multperiod_df):
        lines.append(('', ''))
        lines.append(('各檔排程', ''))
        for _, r in multperiod_df.sort_values(['period', 'y_hat'], ascending=[True, False]).iterrows():
            lines.append(
                (str(r['period']), f"{r['new_song_id']}｜{r['profile_name']}｜ŷ={r['y_hat']:.4f}｜成本={r['cost']:.0f}")
            )
    if comparison is not None and len(comparison):
        lines += [('', ''), ('方法比較摘要', comparison.to_string(index=False))]
    lines += [
        ('', ''),
        ('FLOP–HOP–TOP 前 3（獨立排序）', '、'.join(ranked.head(3)['new_song_id'].tolist())),
    ]
    return lines


def _write_narrative_sheet(ws, rows: list[tuple[str, str]], title: str) -> None:
    from openpyxl.styles import Alignment, Font

    ws['A1'] = title
    ws['A1'].font = Font(bold=True, size=14)
    ws.merge_cells('A1:B1')
    start = 3
    for i, (k, v) in enumerate(rows, start=start):
        ws.cell(row=i, column=1, value=k)
        ws.cell(row=i, column=2, value=v)
        ws.cell(row=i, column=1).font = Font(bold=True) if v == '' and k else Font()
        for col in (1, 2):
            ws.cell(row=i, column=col).alignment = Alignment(wrap_text=True, vertical='top')
    ws.column_dimensions['A'].width = 24
    ws.column_dimensions['B'].width = 78


def export_strategy_workbook(
    *,
    csv_path: Path,
    profile: dict,
    cfg: dict,
    ranked: pd.DataFrame,
    chosen: list[str],
    kparams: dict,
    summary_row: dict,
    baseline: pd.DataFrame,
    sensitivity: pd.DataFrame,
    full_solution: pd.DataFrame,
    multperiod_df: Optional[pd.DataFrame] = None,
    trad_df: Optional[pd.DataFrame] = None,
    comparison: Optional[pd.DataFrame] = None,
    constraint_check: Optional[pd.DataFrame] = None,
    feas_note: Optional[str] = None,
    cv_r2_text: str = '',
    out_path: Path = STRATEGY_EXCEL,
) -> Path:
    try:
        from openpyxl import Workbook
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError as e:
        raise ImportError('請安裝 openpyxl: .venv/bin/python -m pip install openpyxl') from e

    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    wb = Workbook()

    ws_bg = wb.active
    ws_bg.title = '01_背景前提'
    _write_narrative_sheet(ws_bg, _background_rows(csv_path, profile, cfg, cv_r2_text), '背景前提')

    ws_dec = wb.create_sheet('02_決策結論')
    _write_narrative_sheet(
        ws_dec,
        _decision_rows(chosen, kparams, summary_row, ranked, multperiod_df, comparison, feas_note),
        'ILP 決策結論',
    )

    if comparison is not None and len(comparison):
        ws_cmp = wb.create_sheet('03_方法比較')
        ws_cmp.append(['傳統 A&R vs ILP 多限制式背包'])
        for r in dataframe_to_rows(comparison, index=False, header=True):
            ws_cmp.append(r)

    ws_songs = wb.create_sheet('04_候選歌曲評分')
    cols = [
        'rank_by_yhat', 'new_song_id', 'profile_name', 'y_hat', 'tier', 'flop_hop_top',
        'cost', 'selected', 'period', 'loudness', 'tempo', 'danceability', 'energy',
    ]
    show = [c for c in cols if c in full_solution.columns]
    df_show = full_solution[show].sort_values('y_hat', ascending=False)
    ws_songs.append([f'{N_SYNTHETIC_SONGS} 首候選；selected=1 為 ILP 選中'])
    for r in dataframe_to_rows(df_show, index=False, header=True):
        ws_songs.append(r)

    if multperiod_df is not None and len(multperiod_df):
        ws_ilp = wb.create_sheet('05_ILP排程')
        for r in dataframe_to_rows(multperiod_df, index=False, header=True):
            ws_ilp.append(r)

    if trad_df is not None and len(trad_df):
        ws_tr = wb.create_sheet('06_傳統排程')
        for r in dataframe_to_rows(trad_df, index=False, header=True):
            ws_tr.append(r)

    if constraint_check is not None and len(constraint_check):
        ws_con = wb.create_sheet('07_聲學門檻檢核')
        for r in dataframe_to_rows(constraint_check, index=False, header=True):
            ws_con.append(r)

    ws_cost = wb.create_sheet('08_每首成本')
    cost_df = pd.DataFrame([
        {'profile_name': k, 'cost': v} for k, v in COST_BY_PROFILE.items()
    ])
    for r in dataframe_to_rows(cost_df, index=False, header=True):
        ws_cost.append(r)

    ws_sens = wb.create_sheet('09_預算敏感度')
    ws_sens.append(['單檔期總預算敏感度（對照）'])
    for r in dataframe_to_rows(sensitivity, index=False, header=True):
        ws_sens.append(r)

    if SYNTHETIC_CSV.exists():
        ws_syn = wb.create_sheet('10_模擬特徵')
        syn = pd.read_csv(SYNTHETIC_CSV)
        key = [c for c in ['new_song_id', 'profile_name', 'loudness', 'tempo', 'duration', 'year',
                           'danceability', 'energy'] if c in syn.columns]
        for r in dataframe_to_rows(syn[key], index=False, header=True):
            ws_syn.append(r)

    ws_param = wb.create_sheet('11_參數摘要')
    import json
    param_items = {
        '資料檔': str(csv_path),
        '原型列': cfg.get('base_row'),
        '各檔預算': json.dumps(kparams.get('budget_by_period', {}), ensure_ascii=False),
        '各檔 L/U': f"L={kparams.get('l_min_by_period')} U={kparams.get('u_max_by_period')}",
        '聲學門檻': json.dumps(kparams.get('acoustic', {}), ensure_ascii=False),
        'ILP 選中': ','.join(chosen),
    }
    for k, v in param_items.items():
        ws_param.append([k, v])

    wb.save(out_path)
    return out_path
